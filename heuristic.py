# -*- coding:utf-8 -*-
import os
import re
import pandas as pd
import openpyxl
from openpyxl import Workbook
import math
import re
Filename = "bana_raw_dataset.txt"
dict = {}
threshold = 5
# load doc into memory
def load_data(filename):
	# open the file as read only
	file = open(filename, 'r', encoding='utf-8')
	# read all text
	text = file.read()
	# close the file
	file.close()
	return text
two_char = ['c̆', 'ĕ', 'ĭ', 'ŏ', 'ơ̆', 'ŭ', 'ư̆', 'C̆', 'Ĕ', 'Ĭ', 'Ŏ', 'Ơ̆', 'Ŭ', 'Ư̆']
three_char = ['ĕ̂', 'ŏ̂', 'Ĕ̂', 'Ŏ̂']
CHAR_REPLACE = "a b c d e f g h i j k l m n o p q r s t u v w y z A B C D E F G H I J K L M N O P Q R S T U V ạ ả ã à á â ậ ầ ấ ẩ ẫ ă ắ ằ ặ ẳ ẵ ó ò ọ õ ỏ ô ộ ổ ỗ ồ ố ơ ờ ớ ợ ở ỡ é è ẻ ẹ ẽ ê ế ề ệ ể ễ ú ù ụ ủ ũ ư ự ữ ử ừ ứ í ì ị ỉ ĩ ý ỳ ỷ ỵ ỹ đ Ạ Ả Ã À Á Â Ậ Ầ Ấ Ẩ Ẫ Ă Ắ Ằ Ặ Ẳ Ẵ Ó Ò Ọ Õ Ỏ Ô Ộ Ổ Ỗ Ồ Ố Ơ Ờ Ớ Ợ Ở Ỡ É È Ẻ Ẹ Ẽ Ê Ế Ề Ệ Ể Ễ Ú Ù Ụ Ủ Ũ Ư Ự Ữ Ử Ừ Ứ Í Ì Ị Ỉ Ĩ Ý Ỳ Ỷ Ỵ Ỹ Đ a ă â b ƀ c̆ d đ e ĕ ê ĕ̂ g h i ĭ j k l m n ñ o ŏ ô ŏ̂ ơ ơ̆ p r s t u ŭ ư ư̆ w y f q v z A Ă Â B Ƀ C̆ D Đ E Ĕ Ê Ĕ̂ G H I Ĭ J K L M N Ñ O Ŏ Ô Ŏ̂ Ơ Ơ̆ P R S T U Ŭ Ư Ư̆ W Y F Q V Z"
char = CHAR_REPLACE.split(' ')
char_skip = ['"', '”', '(', ')', '_', ',', ';', ':', '.', '/', '\n',' ','>','<']
char_delete = ['@','#','$', '%', '|', '&', '*', '^', "'", "¬", "Œ", "ˆ", "£"]
def split_word(text):
    # result = re.split(r'[\s/()“”]', text)
    word_list = []
    space_list = []
    space_first = -1
    i = 0
    n = len(text)
    current_word = ""
    current_space = ""
    
    while i < n:
        if text[i] in char_delete:
            i+=1
            continue
        if text[i] in char_skip:
            if space_first == -1:
                space_first = True
            if current_word != "":
                word_list.append(current_word)
                current_word = ""
            current_space += text[i]
        else:
            if space_first == -1:
                space_first = False
            if current_space != "":
                space_list.append(current_space)
                current_space = ""
            current_word += text[i]
        i += 1
    if current_word != "":
        word_list.append(current_word)
        current_word = ""
    if current_space != "":
        space_list.append(current_space)
        current_space = ""
    # print(word_list, space_list)
    return word_list, space_list, space_first
# print(split_word("đẩy (ra/vào) [a"))
def split_character(text):
    result = []
    n = len(text)
    i = 0
    while (i < n):
        check = False
        for j in three_char:
            if i+2 < len(text) and text[i] + text[i+1] + text[i+2] == j:
                result.append(j)
                i = i+3
                check = True
                break
        for j in two_char:
            
            if i+1 < len(text) and text[i] + text[i+1] == j:
                # print(j)
                result.append(j)
                i = i+2
                check = True
                break
        if not check:
            result.append(text[i])
            i = i+1
    return result

def text_cleaner(text):
    result = re.sub('[,_“”();:.]',"", text) # TH thay thế là dấu , ' , "", ( )
    result = re.split('\s', result)
    # result = split_word(text)[0] 
    # temp_result = []
    for i in result:
        if len(i) <= 1:
            result.remove(i)
            continue
        # if re.search("^c̆", i):
        #     clone = 'c' + i[2:]
        #     temp_result.append(clone)
        # if re.search("^C̆", i):
        #     clone = 'C' + i[2:]
        #     temp_result.append(clone)
        # if re.search("^ƀ", i):
        #     clone = 'b' + i[1:]
        #     temp_result.append(clone)    
        # if re.search("^Ƀ", i):
        #     clone = 'B' + i[1:]
        #     temp_result.append(clone)
    # for i in temp_result:
    #     result.append(i)
    # print(result)
    for temp in result:
        i = split_character(temp)
        n = len(i)
        # print("length: ", n)
        for j in range(0, len(i)-1):
        
            temp = i[j]+i[j+1]
            if (temp, n) in dict.keys():
                dict[(temp, n)] += 1
            else: 
                dict[(temp, n)] = 1
        for j in range(0, len(i)-2):
            temp = i[j]+i[j+1]+i[j+2]
            if (temp, n) in dict.keys():
                dict[(temp, n)] += 1
            else: 
                dict[(temp, n)] = 1
        for j in range(0, len(i)-3):
            temp = i[j]+i[j+1]+i[j+2]+i[j+3]
            if (temp, n) in dict.keys():
                dict[(temp, n)] += 1
            else: 
                dict[(temp, n)] = 1
        for j in range(0, len(i)-4):
            temp = i[j]+i[j+1]+i[j+2]+i[j+3]+i[j+4]
            if (temp, n) in dict.keys():
                dict[(temp, n)] += 1
            else: 
                dict[(temp, n)] = 1

def findreplacement(text, n):
    
    result = 0
    current = None
    char_list = split_character(text)
    for i in range(len(char_list)): # duyet tren substring bi sai
        for j in char: # j la kí tự để thay thế
            pre_substr = ""
            after_substr = ""
            for l in range(0,i):
                pre_substr += char_list[l]
            for l in range(i+1,len(char_list)):
                after_substr += char_list[l]
            temp = pre_substr + j + after_substr
            if (temp, n) in dict and dict[(temp, n)] >= threshold and dict[(temp, n)] >= result:
                result = dict[(temp, n)]
                current = temp
    # print("Thay thế: ", text, n, current)
    return current # current là từ có xác xuất đúng cao nhất               
def correction(text): # phiên bản hiện tại correct được cho cả câu.
    char_list_raw = split_character(text)
    # print(char_list_raw)
    char_list = []
    char_list_position = []
    for i in char_list_raw:
        if i in char_delete:
            char_list_raw.remove(i)
    for i in range(len(char_list_raw)):
        if char_list_raw[i] == ']' :
            char_list_raw[i] = 'l'
        if char_list_raw[i] == 'š':
            char_list_raw[i] = 'c̆'
        if char_list_raw[i] == '“':
            char_list_raw[i] = '‘'
            # print("YES")    
        if char_list_raw[i] not in char_skip:
            char_list.append(char_list_raw[i])
            char_list_position.append(i)
    length = len(char_list_raw)
    # print(char_list)
    for i in range(len(char_list)):
        for j in reversed(range(2, 6)):
            if i+j > len(char_list):
                continue
            substr = ""
            for k in range(0, j):
                substr = substr + char_list[i+k]
            n = len(char_list)
            # print(i, j, substr, n)
            if (substr, n) not in dict or dict[(substr, n)] < threshold: 
                # print("Bi sai: ", substr)
                sub_result = findreplacement(substr, len(char_list))
                if(not sub_result):
                    continue
                pre_substr = ""
                after_substr = ""
                for l in range(0,i):
                    char_list_raw[char_list_position[l]] = char_list[l] 
                    
                for m in range(0, char_list_position[i]):
                    pre_substr += char_list_raw[m]    
                
                for l in range(i+j,len(char_list)):
                    char_list_raw[char_list_position[l]] = char_list[l] 
                    
                for m in range(char_list_position[i+j-1]+1, length):
                    after_substr += char_list_raw[m]
                
                result = pre_substr + sub_result + after_substr
                # print("Start: ", i, " ; End: ", i+j-1)
                return result
            else:
                # print(substr, dict[(substr, len(char_list))])
                break
    result = ""
    for i in char_list_raw:
        result += i
    return result
def correct_manytime(text):
    length = math.ceil(len(text)/2)
    result = text
    while length > 0:
        result = correction(result)
        length -= 1
    return result
def correction_sentence(text):
    if text is None:
        return ""
    final_result = ""
    split = split_word(text)
    word_list = split[0]
    space_list = split[1]
    space_first = split[2]
    for word in word_list: # tách từ từ câu
        if word == "" or word in char_delete:
            word_list.remove(word)
    cnt = 0
    if space_first and cnt < len(space_list):
        final_result += space_list[cnt]
        cnt += 1
    for i in word_list:
        # print(i, cnt)
        if i in dict and dict[i] >= threshold:
            if final_result == "":
                final_result = i
                if cnt < len(space_list):
                    final_result += space_list[cnt]
                cnt+=1
                continue
            else:
                final_result = final_result + i
                if cnt < len(space_list):
                    final_result += space_list[cnt]
                cnt+=1
                continue
        if final_result == "":
            final_result = correct_manytime(i)
            if cnt < len(space_list):
                final_result += space_list[cnt]
        else:
            final_result = final_result + correct_manytime(i)
            if cnt < len(space_list):
                final_result += space_list[cnt]
        cnt+=1
    while cnt < len(space_list):
        final_result += space_list[cnt]
        cnt += 1
    return final_result
raw_text = load_data(Filename)
text_cleaner(raw_text)

# dataframe1 = openpyxl.load_workbook('.\excel\Tu_dien_Viet_Bahnar.xlsx') #, sheet_name = "VIET - BANA"

# df2 = pd.read_excel('.\excel\Tu_dien_Viet_Bahnar.xlsx', sheet_name = "BANA")
folder = ".\excel"
def correct_excel(file_name, sheet_name):
    wb = openpyxl.load_workbook(folder + "\\" + file_name) # get work book 
    ws = wb[sheet_name] # get work sheet
    result_file_name = "corrected_" + file_name # name of new work book for storing corrected result
    final_wb = openpyxl.load_workbook(folder + "\\" + result_file_name)
    # final_wb.create_sheet(sheet_name)
    final_ws = final_wb[sheet_name]
    num_rows, num_cols = ws.max_row, ws.max_column
    print(num_rows, num_cols)
    # rows1 = ws.iter_rows(min_row = 1, max_row = num_rows, min_col=1, max_col= num_cols)
    for i in range(1, num_rows+1):
        for j in range(1, num_cols+1):
            value = ws.cell(row=i, column = j).value
            # correct_value = correction_sentence(value)
            c1 = final_ws.cell(row = i, column = j)
            
            c1.value = correction_sentence(value) # writing values to cells
            # print(j, value)
            # print(i)
            print(i, "Pre: ", value, " ; After: ", correction_sentence(value))
            # print(j.value)
    final_wb.save(folder + "\\" + result_file_name)
# correct_excel("Tu_dien_Viet_Bahnar.xlsx", "VIET - BANA")
# correct_excel("Tu_dien_Viet_Bahnar.xlsx", "BANA")
correct_excel("Tu_dien_Bahnar_Viet.xlsx", "Sheet1")
# df1.to_excel('corrected_file_Tu_dien_Viet_Bahnar.xlsx, sheet_name = "VIET - BANA"')